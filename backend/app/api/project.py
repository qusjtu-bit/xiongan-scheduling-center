# -*- coding: utf-8 -*-
"""
规建管一体化 API（Step 5）

项目列表/详情、阶段时间轴、审批记录、进度预警、统计分析
"""
from datetime import date, datetime
import io

from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, current_user
from app import db
from app.models import ProjectInfo, ProjectStage, ApprovalRecord, Todo, Message

project_bp = Blueprint('project', __name__)

STAGE_ORDER = ['立项', '规划', '审批', '建设', '验收', '运维']


def _today():
    return date.today().isoformat()


def _parse_date(s):
    """容忍 YYYY-MM 格式，补全为 YYYY-MM-DD"""
    if not s:
        return None
    if len(s) == 7:  # YYYY-MM
        s = s + '-01'
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


# =========================== 项目列表 ===========================

@project_bp.route('/projects', methods=['GET'])
@jwt_required()
def list_projects():
    q = ProjectInfo.query

    # 筛选条件
    keyword = request.args.get('keyword', '').strip()
    ptype = request.args.get('ptype', '').strip()
    area = request.args.get('area', '').strip()
    stage = request.args.get('stage', '').strip()
    alert = request.args.get('alert', '').strip()  # overdue / near_due

    if keyword:
        q = q.filter(ProjectInfo.name.like(f'%{keyword}%'))
    if ptype:
        q = q.filter_by(ptype=ptype)
    if area:
        q = q.filter_by(area=area)
    if stage:
        q = q.filter_by(stage=stage)

    # 预警筛选：逾期或即将到期
    today = _today()
    today_date = date.today()
    if alert == 'overdue':
        # Filter in Python for date-tolerant comparison
        q = q.filter(ProjectInfo.progress < 100)
    elif alert == 'near_due':
        q = q.filter(ProjectInfo.progress < 80)

    items = q.order_by(ProjectInfo.id).all()

    # 判断是否逾期/临近
    result = []
    for p in items:
        d = p.to_dict()
        plan_end = _parse_date(p.plan_end_date)
        d['is_overdue'] = (plan_end is not None and plan_end < today_date and p.progress < 100)
        d['stage_order'] = STAGE_ORDER.index(p.stage) if p.stage in STAGE_ORDER else -1
        # 进度风险
        if p.progress >= 100:
            d['risk'] = 'none'
        elif d['is_overdue']:
            d['risk'] = 'overdue'
        elif plan_end and plan_end <= date(2026, 9, 1) and p.progress < 80:
            d['risk'] = 'near'
        else:
            d['risk'] = 'normal'
        result.append(d)

    return jsonify(code=200, message='success', data={
        'total': len(result),
        'list': result,
    })


# =========================== 项目详情 ===========================

@project_bp.route('/projects/<int:pid>', methods=['GET'])
@jwt_required()
def project_detail(pid):
    p = ProjectInfo.query.get_or_404(pid)

    # 阶段时间轴
    stages = ProjectStage.query.filter_by(project_id=pid).order_by(
        ProjectStage.stage_order).all()

    # 审批记录
    approvals = ApprovalRecord.query.filter_by(project_id=pid).order_by(
        ApprovalRecord.id.desc()).all()

    # 判断预警
    plan_end = _parse_date(p.plan_end_date)
    today_date = date.today()
    overdue = plan_end is not None and plan_end < today_date and p.progress < 100
    risk = 'overdue' if overdue else ('near' if plan_end and plan_end <= date(2026, 9, 1) and p.progress < 80 else 'normal')

    return jsonify(code=200, message='success', data={
        'project': p.to_dict(),
        'stages': [s.to_dict() for s in stages],
        'approvals': [a.to_dict() for a in approvals],
        'risk': risk,
        'is_overdue': overdue,
    })


# =========================== 项目统计 ===========================

@project_bp.route('/projects/stats', methods=['GET'])
@jwt_required()
def project_stats():
    items = ProjectInfo.query.all()
    today_date = date.today()
    sep_first = date(2026, 9, 1)

    # 按阶段统计
    stage_count = {}
    for s in STAGE_ORDER:
        stage_count[s] = 0
    for p in items:
        if p.stage in stage_count:
            stage_count[p.stage] += 1

    # 按类型统计
    type_count = {}
    for p in items:
        t = p.ptype or '其他'
        type_count[t] = type_count.get(t, 0) + 1

    # 按片区统计
    area_count = {}
    for p in items:
        a = p.area or '其他'
        area_count[a] = area_count.get(a, 0) + 1

    # 投资总额
    total_invest = sum(p.invest for p in items)

    # 进度预警
    overdue = 0
    near_due = 0
    for p in items:
        plan = _parse_date(p.plan_end_date)
        if plan and plan < today_date and p.progress < 100:
            overdue += 1
        elif plan and today_date <= plan <= sep_first and p.progress < 80:
            near_due += 1
    normal = len(items) - overdue - near_due

    return jsonify(code=200, message='success', data={
        'total': len(items),
        'total_invest': total_invest,
        'stage_count': stage_count,
        'type_count': type_count,
        'area_count': area_count,
        'risk_summary': {
            'overdue': overdue,
            'near_due': near_due,
            'normal': normal,
        },
    })


# =========================== 预警列表 ===========================

@project_bp.route('/projects/alerts', methods=['GET'])
@jwt_required()
def project_alerts():
    items = ProjectInfo.query.all()
    today_date = date.today()
    sep_first = date(2026, 9, 1)
    alerts = []

    for p in items:
        plan_end = _parse_date(p.plan_end_date)
        if not plan_end:
            continue
        if plan_end < today_date and p.progress < 100:
            days = (today_date - plan_end).days
            alerts.append({
                'project_id': p.id,
                'project_name': p.name,
                'alert_type': 'overdue',
                'alert_label': '逾期',
                'days': days,
                'plan_end': p.plan_end_date,
                'progress': p.progress,
                'stage': p.stage,
            })
        elif today_date <= plan_end <= sep_first and p.progress < 80:
            alerts.append({
                'project_id': p.id,
                'project_name': p.name,
                'alert_type': 'near_due',
                'alert_label': '临近',
                'plan_end': p.plan_end_date,
                'progress': p.progress,
                'stage': p.stage,
                'days': None,
            })

    # 逾期排最前
    alerts.sort(key=lambda x: (0 if x['alert_type'] == 'overdue' else 1, x.get('plan_end', '')))

    return jsonify(code=200, message='success', data={
        'total': len(alerts),
        'list': alerts,
    })


# =========================== 待审批列表 ===========================

@project_bp.route('/projects/pending-approvals', methods=['GET'])
@jwt_required()
def pending_approvals():
    items = ApprovalRecord.query.filter_by(status='待审批').order_by(
        ApprovalRecord.apply_date.asc()).all()
    today = date.today()
    result = []
    for a in items:
        p = ProjectInfo.query.get(a.project_id)
        d = a.to_dict()
        d['project_name'] = p.name if p else ''
        # 超时判断：申请超过7天未处理
        d['is_overdue'] = False
        if a.apply_date:
            d1 = _parse_date(a.apply_date)
            if d1 and (today - d1).days > 7:
                d['is_overdue'] = True
        result.append(d)
    return jsonify(code=200, message='success', data={'total': len(result), 'list': result})


# =========================== 审批操作（实务闭环） ===========================

@project_bp.route('/projects/approvals/<int:aid>/action', methods=['POST'])
@jwt_required()
def approval_action(aid):
    """审批通过/驳回操作"""
    a = ApprovalRecord.query.get_or_404(aid)
    if a.status != '待审批':
        return jsonify(code=400, message='该审批已处理，请勿重复操作'), 400

    d = request.get_json(silent=True) or {}
    action = d.get('action', '')  # 'approve' | 'reject'
    comment = d.get('comment', '').strip()

    if action not in ('approve', 'reject'):
        return jsonify(code=400, message='操作类型仅支持 approve/reject'), 400

    now_date = date.today().isoformat()
    operator = (current_user.real_name or current_user.username)
    operator_dept = (current_user.dept.dept_name if current_user.dept else '')

    if action == 'approve':
        a.status = '已通过'
        a.approve_date = now_date
        a.remark = (a.remark or '') + f' | {operator}({operator_dept}) 批准于 {now_date}'
        if comment:
            a.remark += f' · 批注：{comment}'
    else:
        a.status = '已驳回'
        a.remark = (a.remark or '') + f' | {operator}({operator_dept}) 驳回于 {now_date}'
        if comment:
            a.remark += f' · 原因：{comment}'

    db.session.commit()

    # 发送消息通知（局级可见）
    p = ProjectInfo.query.get(a.project_id)
    if p:
        msg_title = f'审批{"通过" if action == "approve" else "驳回"}：{a.approval_type}'
        msg_content = f'项目「{p.name}」的 {a.approval_type} 已被 {operator}({operator_dept}) {"批准" if action == "approve" else "驳回"}。{comment if comment else ""}'
        msg = Message(
            sender=operator,
            title=msg_title,
            content=msg_content,
            msg_type=2,
            level=2,
            scope=1,
            scope_id=0,
            read_users='',
        )
        db.session.add(msg)
        db.session.commit()

    return jsonify(code=200, message='操作成功', data={
        'id': a.id,
        'status': a.status,
        'approve_date': a.approve_date,
        'remark': a.remark,
        'operator': operator,
        'operator_dept': operator_dept,
    })


# =========================== 专题工作台（业务实用性改造） ===========================

# 五大业务线 → 涉及处室映射
TOPIC_DEPT_MAP = {
    1: ['政务服务处', '工程质量安全处', '城乡发展处', '城市建设监察处', '房屋管理处'],
    2: ['建筑市场处', '工程质量安全处', '城市建设监察处', '政策法规处'],
    3: ['水利组', '城市管理处', '综合交通组'],
    4: ['城乡发展处', '工程质量安全处', '信息化处'],
    5: ['城市管理处', '信息化处', '房屋管理处', '综合交通组', '水利组'],
}

# 五大业务线 → 关联领域
TOPIC_DOMAIN_MAP = {1: 1, 2: 1, 3: 3, 4: 1, 5: 4}


@project_bp.route('/topic-workbench/<int:topic_id>', methods=['GET'])
@jwt_required()
def topic_workbench(topic_id):
    """业务专题工作台——以'干活'为导向，非纯展示"""
    if topic_id not in TOPIC_DEPT_MAP:
        return jsonify(code=404, message=f'专题 {topic_id} 不存在'), 404

    depts = TOPIC_DEPT_MAP[topic_id]
    domain = TOPIC_DOMAIN_MAP.get(topic_id, 1)

    # 1. 本业务线关联的待办事项（从 Todo）
    todos = Todo.query.filter(
        Todo.status.in_([0, 1])  # 待办/处理中
    ).order_by(Todo.urgency.desc(), Todo.due_date.asc()).limit(10).all()
    todo_list = []
    for t in todos:
        td = t.to_dict()
        # 简化判断：如果待办内容提到专题相关处室
        match = any(d in (t.title or '') for d in depts)
        if match:
            todo_list.append(td)

    # 2. 本业务线关联的审批记录
    approvals = ApprovalRecord.query.filter(
        ApprovalRecord.status == '待审批'
    ).order_by(ApprovalRecord.apply_date.asc()).all()
    pending_approvals = []
    for a in approvals:
        p = ProjectInfo.query.get(a.project_id)
        if not p:
            continue
        # 通过项目阶段关联处室判断是否属于本业务线
        stages = ProjectStage.query.filter_by(project_id=a.project_id).all()
        matched = any(any(d in (s.resp_dept or '') for d in depts) for s in stages)
        if matched:
            ad = a.to_dict()
            ad['project_name'] = p.name
            pending_approvals.append(ad)

    # 3. 本业务线的项目预警
    items = ProjectInfo.query.all()
    today_date = date.today()
    alert_list = []
    for p in items:
        if p.progress >= 100:
            continue
        plan_end = _parse_date(p.plan_end_date)
        if not plan_end:
            continue
        # 检查项目阶段是否有本业务线处室参与
        stages = ProjectStage.query.filter_by(project_id=p.id).all()
        involved = any(any(d in (s.resp_dept or '') for d in depts) for s in stages)
        if not involved:
            continue
        if plan_end < today_date:
            alert_list.append({
                'project_id': p.id, 'project_name': p.name,
                'alert_type': 'overdue', 'days': (today_date - plan_end).days,
                'stage': p.stage, 'progress': p.progress,
            })
        elif plan_end <= date(2026, 9, 1) and p.progress < 80:
            alert_list.append({
                'project_id': p.id, 'project_name': p.name,
                'alert_type': 'near_due', 'stage': p.stage, 'progress': p.progress,
                'days': (plan_end - today_date).days,
            })

    # 按紧急程度排序
    alert_list.sort(key=lambda x: (0 if x['alert_type'] == 'overdue' else 1, -(x.get('days') or 0)))

    # 4. 跨处室流转统计（各阶段审批在哪些处室停留）
    flow_stats = {}
    for s in ProjectStage.query.all():
        if not s.resp_dept:
            continue
        for d in s.resp_dept.split(','):
            d = d.strip()
            if d in depts:
                flow_stats[d] = flow_stats.get(d, 0) + 1

    return jsonify(code=200, message='success', data={
        'topic_id': topic_id,
        'departments': depts,
        'todos': todo_list[:5],
        'todo_count': len(todo_list),
        'pending_approvals': pending_approvals[:5],
        'approval_count': len(pending_approvals),
        'alerts': alert_list[:5],
        'alert_count': len(alert_list),
        'flow_stats': [{'dept': k, 'items': v} for k, v in sorted(flow_stats.items())],
    })


# =========================== Excel 台账导出 ===========================

@project_bp.route('/projects/export', methods=['GET'])
@jwt_required()
def export_projects():
    """导出项目台账为 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    items = ProjectInfo.query.order_by(ProjectInfo.id).all()
    wb = Workbook()

    # === Sheet1: 项目台账 ===
    ws = wb.active
    ws.title = '项目台账'
    headers = ['序号', '项目名称', '类型', '片区', '建设单位', '施工单位', '监理单位',
               '投资额(万元)', '建设规模', '当前阶��', '进度(%)', '开工日期', '计划竣工', '风险状态']
    ws.append(headers)

    # 表头样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    today_date = date.today()
    for i, p in enumerate(items, 1):
        plan_end = _parse_date(p.plan_end_date)
        risk = '正常'
        if p.progress >= 100:
            risk = '已完成'
        elif plan_end and plan_end < today_date:
            risk = '逾期'
        elif plan_end and plan_end <= date(2026, 9, 1) and p.progress < 80:
            risk = '临近'

        ws.append([
            i, p.name, p.ptype, p.area, p.build_unit, p.contractor or '',
            p.supervisor or '', round(p.invest, 1), f'{p.scale} {p.scale_unit}',
            p.stage, p.progress, p.start_date or '', p.plan_end_date or '', risk,
        ])

    # 列宽
    widths = [6, 28, 8, 10, 20, 20, 20, 12, 14, 10, 10, 12, 12, 8]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else 'A'].width = w

    # === Sheet2: 审批记录台账 ===
    ws2 = wb.create_sheet('审批记录')
    appr_headers = ['序号', '项目名称', '审批类型', '申请日期', '审批日期', '状态', '审批人', '备注']
    ws2.append(appr_headers)
    for col in range(1, len(appr_headers) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    approvals = ApprovalRecord.query.order_by(ApprovalRecord.id).all()
    for i, a in enumerate(approvals, 1):
        p = ProjectInfo.query.get(a.project_id)
        ws2.append([
            i, p.name if p else '', a.approval_type, a.apply_date or '',
            a.approve_date or '', a.status, a.approver or '', a.remark or '',
        ])

    appr_widths = [6, 28, 16, 12, 12, 8, 20, 40]
    for col, w in enumerate(appr_widths, 1):
        ws2.column_dimensions[chr(64 + col) if col <= 26 else 'A'].width = w

    # === Sheet3: 阶段台账 ===
    ws3 = wb.create_sheet('阶段里程碑')
    stage_headers = ['序号', '项目名称', '阶段名称', '顺序', '开始日期', '计划完成', '实际完成', '状态', '责任处室', '备注']
    ws3.append(stage_headers)
    for col in range(1, len(stage_headers) + 1):
        cell = ws3.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    stages = ProjectStage.query.order_by(ProjectStage.project_id, ProjectStage.stage_order).all()
    for i, s in enumerate(stages, 1):
        p = ProjectInfo.query.get(s.project_id)
        ws3.append([
            i, p.name if p else '', s.stage_name, s.stage_order,
            s.start_date or '', s.plan_end_date or '', s.actual_end_date or '',
            s.status, s.resp_dept or '', s.remark or '',
        ])

    # 导出
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'建交局项目台账_{date.today().isoformat()}.xlsx'
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


# =========================== 审批时效看板 ===========================

@project_bp.route('/approvals/efficiency', methods=['GET'])
@jwt_required()
def approval_efficiency():
    """审批时效分析——按处室统计平均审批时长、超时率、待办积压"""
    from collections import defaultdict

    all_approvals = ApprovalRecord.query.all()
    today_str = date.today().isoformat()

    # 按审批人(处室)分组统计
    dept_stats = defaultdict(lambda: {
        'total': 0, 'approved': 0, 'rejected': 0, 'pending': 0,
        'processing_days': [], 'overdue_count': 0,
    })

    for a in all_approvals:
        approvers = (a.approver or '').split(',')
        for dept in approvers:
            dept = dept.strip()
            if not dept:
                continue
            s = dept_stats[dept]
            s['total'] += 1
            if a.status == '已通过':
                s['approved'] += 1
                # 计算审批时长
                if a.apply_date and a.approve_date:
                    try:
                        d1 = _parse_date(a.apply_date)
                        d2 = _parse_date(a.approve_date)
                        if d1 and d2:
                            s['processing_days'].append((d2 - d1).days)
                    except Exception:
                        pass
            elif a.status == '已驳回':
                s['rejected'] += 1
            elif a.status == '待审批':
                s['pending'] += 1
                # 判断是否超时（申请超过7天未处理）
                if a.apply_date:
                    try:
                        d1 = _parse_date(a.apply_date)
                        if d1 and (date.today() - d1).days > 7:
                            s['overdue_count'] += 1
                    except Exception:
                        pass

    # 构建结果
    result = []
    for dept, s in sorted(dept_stats.items(), key=lambda x: -x[1]['pending']):
        avg_days = round(sum(s['processing_days']) / len(s['processing_days']), 1) if s['processing_days'] else 0
        on_time_rate = round((s['approved'] + s['rejected']) / s['total'] * 100, 1) if s['total'] else 0
        result.append({
            'dept': dept,
            'total': s['total'],
            'approved': s['approved'],
            'rejected': s['rejected'],
            'pending': s['pending'],
            'overdue': s['overdue_count'],
            'avg_days': avg_days,
            'on_time_rate': on_time_rate,
            'efficiency_label': '高效' if avg_days <= 2 and s['overdue_count'] == 0
                              else '正常' if avg_days <= 5
                              else '待改善',
        })

    # 全局统计
    total_all = sum(s['total'] for s in dept_stats.values())
    pending_all = sum(s['pending'] for s in dept_stats.values())
    overdue_all = sum(s['overdue_count'] for s in dept_stats.values())
    all_days = []
    for s in dept_stats.values():
        all_days.extend(s['processing_days'])
    avg_all = round(sum(all_days) / len(all_days), 1) if all_days else 0

    return jsonify(code=200, message='success', data={
        'departments': result,
        'summary': {
            'total_approvals': total_all,
            'pending': pending_all,
            'overdue': overdue_all,
            'avg_days': avg_all,
            'on_time_rate': round((total_all - overdue_all) / total_all * 100, 1) if total_all else 100,
        },
    })


# =========================== 批量审批 ===========================

@project_bp.route('/approvals/batch-action', methods=['POST'])
@jwt_required()
def batch_approval_action():
    """批量审批通过/驳回"""
    d = request.get_json(silent=True) or {}
    action = d.get('action', '')
    ids = d.get('ids', [])
    comment = d.get('comment', '').strip()

    if action not in ('approve', 'reject'):
        return jsonify(code=400, message='操作类型仅支持 approve/reject'), 400
    if not ids or not isinstance(ids, list):
        return jsonify(code=400, message='请选择至少一条审批'), 400

    now_date = date.today().isoformat()
    operator = (current_user.real_name or current_user.username)
    operator_dept = (current_user.dept.dept_name if current_user.dept else '')
    labels = {'approve': '批准', 'reject': '驳回'}

    success = 0
    skipped = 0
    for aid in ids:
        a = ApprovalRecord.query.get(aid)
        if not a or a.status != '待审批':
            skipped += 1
            continue
        if action == 'approve':
            a.status = '已通过'
            a.approve_date = now_date
            a.remark = (a.remark or '') + f' | {operator}({operator_dept}) 批量批准于 {now_date}'
        else:
            a.status = '已驳回'
            a.remark = (a.remark or '') + f' | {operator}({operator_dept}) 批量驳回于 {now_date}'
        if comment:
            a.remark += f' · {comment}'
        success += 1

    db.session.commit()

    # 发送消息
    msg = Message(
        sender=operator,
        title=f'批量审批{labels[action]}：{success}条',
        content=f'{operator}({operator_dept}) 批量{labels[action]}了 {success} 条审批。{comment if comment else ""}',
        msg_type=2, level=2, scope=1, scope_id=0, read_users='',
    )
    db.session.add(msg)
    db.session.commit()

    return jsonify(code=200, message=f'批量操作完成：成功{success}条，跳过{skipped}条', data={
        'success': success, 'skipped': skipped,
    })
